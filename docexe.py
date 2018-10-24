from openpyxl import load_workbook
import xlwt


def get_openpyxl():
    wb = load_workbook('/Users/justin/Documents/python/ExcelSortingJH/files/test_4.xlsx')
    sheet = wb['test4']

    items_to_sort = {'items': []}

    for row in sheet:
        items_to_sort['items'].append((row[0].value, row[1].value, row[2].value))

    return items_to_sort


def write_to_file(data):
    book = xlwt.Workbook(encoding="utf-8")
    sheet2 = book.add_sheet("SORTED rules_pattern_counts(11)")
    ordered_array = []
    insert_data_to_sheet(data, ordered_array, 1)

    for items in ordered_array:
        print("this is a tuple: %s" % (items,))

    count = 0
    for row in ordered_array:
        sheet2.write(count, row['layer'], row['target_freq'])
        sheet2.write(count, (row['layer']+1), row['target'])
        sheet2.write(count, (row['layer'] + 2), row['cluster_count'])

        count += 1
    book.save("Justin_Test_3_copy.xls")


def insert_data_to_sheet(dictionary, ordered_array, layer):
    ordered_array.append({'target_freq': dictionary['target_freq'], 'target': dictionary['target'], 'cluster_count':dictionary['cluster_count'], 'layer': layer})
    for item in dictionary['target_list']:
        insert_data_to_sheet(item, ordered_array, (layer+1))


def read_file():
    get_openpyxl()
    write_to_file()


